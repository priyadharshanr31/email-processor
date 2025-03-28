import pandas as pd
import asyncio
from openpyxl import load_workbook
from config import TEMPLATE_FILE, OUTPUT_FOLDER
import re

async def process_latest_file(file_path, start_row=10):
    """Asynchronously process the latest statement file and update the template with extracted data."""
    try:
        print(f"[INFO] Processing file: {file_path}")

        # Read statement file asynchronously
        df = await asyncio.to_thread(pd.read_excel, file_path, engine="openpyxl")

        if "Status" not in df.columns:
            print("❌ [ERROR] 'Status' column not found in the file.")
            return None, None
        
        df_filtered = df[df["Status"].astype(str).str.upper().isin(["Y", "YES", "DONE"])].copy()

        if df_filtered.empty:
            print("[INFO] No matching data found, skipping processing.")
            return None, None

        # Extract currency values from the "Description" column
        if "Description" in df_filtered.columns:
            df_filtered["Extracted_Desc"] = df_filtered["Description"].apply(
                lambda x: re.search(r"([A-Z]{3} \d+)", str(x)).group(1) if pd.notna(x) and re.search(r"([A-Z]{3} \d+)", str(x)) else None
            )

        # Load template file
        wb = await asyncio.to_thread(load_workbook, TEMPLATE_FILE, keep_vba=True)
        ws = wb["CS"]

        # Read existing data
        existing_data = list(ws.values)
        header_row = list(existing_data[0]) if existing_data else []  # ✅ Convert tuple to list

        # Ensure required columns exist
        for col in ["Extracted_Desc", "Matched Currency"]:
            if col not in header_row:
                header_row.append(col)

        # Add "Matched Currency" column based on template currency
        if "Currency" in df_filtered.columns and "Currency" in header_row:
            df_filtered["Matched Currency"] = df_filtered["Currency"].apply(
                lambda x: "Yes" if x in [row[header_row.index("Currency")] for row in existing_data[1:]] else "No"
            )
        else:
            df_filtered["Matched Currency"] = "No"

        # Ensure data order matches template
        for col in header_row:
            if col not in df_filtered.columns:
                df_filtered[col] = None
        df_filtered = df_filtered[header_row]

        # Format account number columns to remove commas
        account_cols = [col for col in df_filtered.columns if "Account" in col or "Acct" in col]
        for col in account_cols:
            df_filtered[col] = df_filtered[col].astype(str).str.replace(",", "", regex=True)

        # Preserve rows before the replacement
        before_insert = existing_data[:start_row - 1]
        
        # Preserve rows after the replacement
        after_insert = existing_data[start_row - 1 + len(df_filtered):]

        # Convert extracted data to lists for different previews
        extracted_data_preview = df_filtered.copy()
        extracted_data_preview["Description"] = df.loc[df_filtered.index, "Description"]  # Restore "Description"
        extracted_data_preview.drop(columns=["Matched Currency", "Extracted_Desc"], errors="ignore", inplace=True)  # ❌ Removed "Extracted_Desc"

        # Remove commas in Account Number for Preview 1
        for col in account_cols:
            extracted_data_preview[col] = extracted_data_preview[col].astype(str).str.replace(",", "", regex=True)

        extracted_data_list = [list(extracted_data_preview.columns)] + extracted_data_preview.values.tolist()  # ✅ Preview 1 (without "Extracted_Desc")

        # 🔹 Fix: Remove commas from Account Number for **Preview 2**  
        preview_2_final_output = df_filtered.copy()
        for col in account_cols:
            preview_2_final_output[col] = preview_2_final_output[col].astype(str).str.replace(",", "", regex=True)

        final_output_list = [list(preview_2_final_output.columns)] + preview_2_final_output.values.tolist()  # ✅ Preview 2 (now without commas)

        # Clear only the rows that will be replaced
        for _ in range(len(df_filtered)):
            ws.delete_rows(start_row)

        # Rewrite the worksheet
        ws.delete_rows(1, ws.max_row)  # Clear existing data
        ws.append(header_row)  # Add headers

        # Insert data before, extracted, and after
        for row in before_insert[1:]:
            ws.append(row)
        for row in preview_2_final_output.values.tolist():  # ✅ Includes "Matched Currency" and no commas
            ws.append(row)
        for row in after_insert:
            ws.append(row)

        # Save processed file
        output_file = f"{OUTPUT_FOLDER}/Processed_Statement.xlsm"
        await asyncio.to_thread(wb.save, output_file)

        print(f"[✅ SUCCESS] Processed file saved: {output_file}")
        return output_file, extracted_data_list  

    except Exception as e:
        print(f"❌ [ERROR] {e}")
        return None, None
